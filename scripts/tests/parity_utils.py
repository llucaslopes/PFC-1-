"""Utilitarios de comparacao de saidas cientificas (paridade bit-a-bit).

Estes helpers garantem que a refatoracao dos scripts Python de pos-processamento
nao altere uma virgula dos entregaveis do TCC nem dos graficos do artigo.

Tres niveis de comparacao sao oferecidos:

1. Textuais (CSV, MD, SVG, MMD, JSON): diff exato por bytes.
2. PNG: comparacao por assinatura (largura, altura, modo, sha256 dos pixels
   brutos via Pillow). Robusto contra metadados de PNG que podem mudar entre
   execucoes (timestamps embutidos, etc.).
3. XLSX: comparacao por valores de celula (sheet.values do openpyxl), nao por
   bytes do zip, porque XLSX embute timestamps de criacao no manifesto.
"""

from __future__ import annotations

import difflib
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

TEXT_EXTENSIONS = {".csv", ".md", ".svg", ".mmd", ".json", ".txt"}
PNG_EXTENSIONS = {".png"}
XLSX_EXTENSIONS = {".xlsx"}

# Arquivos cuja geracao depende de servico externo (mermaid.ink) e que portanto
# podem nao existir em uma regeneracao offline. Comparamos quando existem mas
# nao falhamos quando ausentes.
OPTIONAL_PATH_PATTERNS = (
    re.compile(r"_mermaid\.(png|svg)$"),
)

# Matplotlib insere a hora atual em <dc:date>...</dc:date> e usa hashes pseudo-
# aleatorios para IDs de <clipPath>, <path id="mXXXX">, etc. Como o conteudo
# visual e identico, normalizamos esses tokens antes do diff.
_SVG_NORMALIZERS = (
    (re.compile(r"<dc:date>[^<]*</dc:date>"), "<dc:date>NORMALIZED</dc:date>"),
    (re.compile(r'(<clipPath id=")p[0-9a-f]{10}(")'), r"\1pNORMALIZED\2"),
    (re.compile(r'(clip-path="url\(#)p[0-9a-f]{10}(\)")'), r"\1pNORMALIZED\2"),
    (re.compile(r'(<path id=")m[0-9a-f]{10}(")'), r"\1mNORMALIZED\2"),
    (re.compile(r'(xlink:href="#)m[0-9a-f]{10}(")'), r"\1mNORMALIZED\2"),
    (re.compile(r'(<use [^>]*xlink:href="#)m[0-9a-f]{10}(")'), r"\1mNORMALIZED\2"),
)

# O README.md em resultados/figuras_tcc/ reporta quantos diagramas Mermaid
# foram efetivamente renderizados (depende de internet). Normalizamos a
# contagem para nao gerar falso-positivo offline.
_README_TCC_NORMALIZERS = (
    (re.compile(r"SVG=\d+/\d+,\s*PNG=\d+/\d+"), "SVG=N/M, PNG=N/M"),
)


def sha256_bytes(data: bytes) -> str:
    """Hash SHA256 hex de um buffer."""
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    """Hash SHA256 hex do conteudo bruto de um arquivo."""
    return sha256_bytes(path.read_bytes())


def png_signature(path: Path) -> dict[str, Any]:
    """Assinatura estavel de um PNG: dimensoes + hash dos pixels.

    Importa Pillow lazy para nao quebrar em ambientes sem PIL (os testes
    textuais funcionam mesmo assim).
    """
    from PIL import Image

    with Image.open(path) as image:
        image.load()
        width, height = image.size
        mode = image.mode
        pixel_bytes = image.tobytes()

    return {
        "width": width,
        "height": height,
        "mode": mode,
        "pixels_sha256": sha256_bytes(pixel_bytes),
        "byte_size": path.stat().st_size,
    }


def xlsx_values(path: Path) -> dict[str, Any]:
    """Assinatura estavel de um XLSX: valores de cada celula em cada sheet.

    Ignora estilos e metadados que variam entre execucoes.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheets[sheet_name] = [list(row) for row in sheet.values]
    workbook.close()

    return {
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
    }


@dataclass
class FileDiff:
    """Resultado de uma comparacao. `ok=True` significa paridade preservada."""

    path: str
    ok: bool
    reason: str | None = None
    detail: str | None = None

    def format(self) -> str:
        if self.ok:
            return f"OK    {self.path}"
        head = f"DIFF  {self.path}: {self.reason}"
        if self.detail:
            return f"{head}\n{self.detail}"
        return head


def is_optional_path(relative: str) -> bool:
    """Retorna True para arquivos cuja ausencia nao causa falha do teste.

    Hoje cobre apenas os diagramas Mermaid (`*_mermaid.png` e `*_mermaid.svg`),
    que dependem do servico externo `mermaid.ink`.
    """
    return any(pattern.search(relative) for pattern in OPTIONAL_PATH_PATTERNS)


def normalize_svg(text: str) -> str:
    """Remove tokens nao-deterministicos de SVG gerado por matplotlib.

    Aplica normalizacoes para `<dc:date>`, IDs de `<clipPath>`, IDs de
    `<path id="mXXXX">` e respectivas referencias `xlink:href`/`clip-path`.
    """
    for pattern, replacement in _SVG_NORMALIZERS:
        text = pattern.sub(replacement, text)
    return text


def normalize_readme_tcc(text: str) -> str:
    """Normaliza a contagem de mermaid no README.md do pacote do TCC."""
    for pattern, replacement in _README_TCC_NORMALIZERS:
        text = pattern.sub(replacement, text)
    return text


def _pick_text_normalizer(relative: str):
    """Escolhe a funcao de normalizacao baseada no caminho relativo do arquivo."""
    rel_lower = relative.lower().replace("\\", "/")
    if rel_lower.endswith(".svg"):
        return normalize_svg
    if rel_lower.endswith("resultados/figuras_tcc/readme.md"):
        return normalize_readme_tcc
    return None


def compare_text_files(
    expected: Path,
    actual: Path,
    *,
    max_diff_lines: int = 40,
    relative: str | None = None,
) -> FileDiff:
    """Compara dois arquivos texto, com normalizacao por tipo quando aplicavel.

    Para SVGs do matplotlib e o README do pacote TCC, aplica normalizacoes
    que removem tokens nao-deterministicos (timestamps, IDs hash). Demais
    arquivos textuais (CSV, MD, MMD, JSON) sao comparados byte-a-byte.

    `relative` opcionalmente identifica o caminho relativo do arquivo (usado
    para escolher o normalizador correto e para o cabecalho do diff). Se
    ausente, usa o nome do arquivo.
    """
    rel = relative if relative is not None else expected.name
    if not actual.exists():
        return FileDiff(rel, ok=False, reason="actual file is missing")
    expected_bytes = expected.read_bytes()
    actual_bytes = actual.read_bytes()
    if expected_bytes == actual_bytes:
        return FileDiff(rel, ok=True)

    try:
        expected_text = expected_bytes.decode("utf-8")
        actual_text = actual_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return FileDiff(
            rel,
            ok=False,
            reason="binary diff (cannot decode as utf-8)",
            detail=f"expected sha256={sha256_bytes(expected_bytes)}, actual sha256={sha256_bytes(actual_bytes)}",
        )

    normalizer = _pick_text_normalizer(rel)
    if normalizer is not None:
        if normalizer(expected_text) == normalizer(actual_text):
            return FileDiff(rel, ok=True)

    expected_lines = expected_text.splitlines(keepends=True)
    actual_lines = actual_text.splitlines(keepends=True)
    diff = list(
        difflib.unified_diff(
            expected_lines,
            actual_lines,
            fromfile=f"baseline/{rel}",
            tofile=f"actual/{rel}",
            n=2,
        )
    )
    detail = "".join(diff[:max_diff_lines])
    if len(diff) > max_diff_lines:
        detail += f"\n... (+{len(diff) - max_diff_lines} more diff lines)"
    reason = "text content differs (after normalization)" if normalizer else "text content differs"
    return FileDiff(rel, ok=False, reason=reason, detail=detail)


def compare_png_signatures(
    expected_sig: dict[str, Any],
    actual_path: Path,
    *,
    rel_name: str,
) -> FileDiff:
    """Compara um PNG atual contra uma assinatura previamente capturada.

    Falha se dimensoes/mode mudarem ou se o sha256 dos pixels divergir.
    Bytes do PNG no disco podem variar (metadados, compressao) sem que isso
    seja problema, mas pixels precisam ser identicos.
    """
    if not actual_path.exists():
        return FileDiff(rel_name, ok=False, reason="actual PNG is missing")
    actual_sig = png_signature(actual_path)
    for key in ("width", "height", "mode"):
        if expected_sig[key] != actual_sig[key]:
            return FileDiff(
                rel_name,
                ok=False,
                reason=f"png {key} differs",
                detail=f"expected {expected_sig[key]} vs actual {actual_sig[key]}",
            )
    if expected_sig["pixels_sha256"] != actual_sig["pixels_sha256"]:
        return FileDiff(
            rel_name,
            ok=False,
            reason="png pixel content differs",
            detail=(
                f"expected pixels_sha256={expected_sig['pixels_sha256']}, "
                f"actual pixels_sha256={actual_sig['pixels_sha256']}"
            ),
        )
    return FileDiff(rel_name, ok=True)


def compare_xlsx_values(expected_values: dict[str, Any], actual_path: Path, *, rel_name: str) -> FileDiff:
    """Compara valores de celula de um XLSX atual contra um snapshot."""
    if not actual_path.exists():
        return FileDiff(rel_name, ok=False, reason="actual XLSX is missing")
    actual_values = xlsx_values(actual_path)
    if expected_values["sheet_names"] != actual_values["sheet_names"]:
        return FileDiff(
            rel_name,
            ok=False,
            reason="xlsx sheet names differ",
            detail=f"expected {expected_values['sheet_names']} vs actual {actual_values['sheet_names']}",
        )
    diffs: list[str] = []
    for sheet_name in expected_values["sheet_names"]:
        expected_rows = expected_values["sheets"][sheet_name]
        actual_rows = actual_values["sheets"][sheet_name]
        if expected_rows == actual_rows:
            continue
        diffs.append(f"  sheet {sheet_name!r}: expected {len(expected_rows)} rows, actual {len(actual_rows)} rows")
        max_rows = max(len(expected_rows), len(actual_rows))
        for index in range(max_rows):
            expected_row = expected_rows[index] if index < len(expected_rows) else None
            actual_row = actual_rows[index] if index < len(actual_rows) else None
            if expected_row != actual_row:
                diffs.append(f"    row {index}: expected {expected_row!r}")
                diffs.append(f"               actual   {actual_row!r}")
                if len(diffs) > 30:
                    diffs.append("    ... (truncated)")
                    break
        if len(diffs) > 30:
            break
    if diffs:
        return FileDiff(rel_name, ok=False, reason="xlsx values differ", detail="\n".join(diffs))
    return FileDiff(rel_name, ok=True)


def classify_extension(path: Path) -> str:
    """Devolve `text`, `png`, `xlsx` ou `other` para o tipo de comparacao a usar."""
    ext = path.suffix.lower()
    if ext in TEXT_EXTENSIONS:
        return "text"
    if ext in PNG_EXTENSIONS:
        return "png"
    if ext in XLSX_EXTENSIONS:
        return "xlsx"
    return "other"


def load_manifest(manifest_path: Path) -> dict[str, Any]:
    """Carrega o manifest.json gerado por `snapshot_baselines.py`."""
    with manifest_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def dump_manifest(manifest: dict[str, Any], manifest_path: Path) -> None:
    """Escreve o manifest.json com formatacao deterministica."""
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    with manifest_path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(manifest, handle, indent=2, sort_keys=True, ensure_ascii=False)
        handle.write("\n")
